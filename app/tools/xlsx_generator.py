"""
文件: tools/xlsx_generator.py | XLSX 文件生成工具
职责: 将结构化数据(表头+行)转换为格式化的 Excel 文件

使用的库: openpyxl
选择理由: 读写兼备（xlsxwriter 只能写不能读）
权衡: openpyxl 写大文件时内存占用高，但本项目数据量小，不影响
"""

import os
from datetime import datetime

# openpyxl 是 Python 操作 Excel .xlsx 文件的标准库
from openpyxl import Workbook                    # 工作簿(整个 Excel 文件)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # 单元格样式
from openpyxl.utils import get_column_letter     # 数字列号转字母(A,B,C...)

from app.config import OUTPUT_DIR               # 输出目录配置
from .registry import registry                   # 工具注册中心


def generate_xlsx(tables: list[dict], output_path: str = "") -> str:
    """
    根据表格数据生成格式化的 XLSX 文件

    参数:
      tables: 表格数据列表，每项包含:
        - name: 工作表名称(sheet name)
        - headers: 列标题数组
        - rows: 数据行数组(每行是字符串数组)
      output_path: 指定输出路径(可选)，不指定则自动生成

    返回:
      生成的 .xlsx 文件绝对路径

    样式设计:
      - 表头: 深蓝色背景 + 白色粗体字 + 居中
      - 数据行: 斑马纹(隔行变色) + 细边框
      - 列宽: 自动适配内容
      - 数字: 自动从字符串转为数值(便于 Excel 计算)
    """
    # 自动生成输出路径（未指定时）
    if not output_path:
        os.makedirs(OUTPUT_DIR, exist_ok=True)          # 确保 output 目录存在
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")    # 时间戳，避免文件名冲突
        output_path = str(OUTPUT_DIR / f"extracted_data_{ts}.xlsx")

    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # 创建工作簿
    wb = Workbook()
    # 删除默认的空白工作表（后面按数据量建表）
    wb.remove(wb.active)

    # 无数据时的占位
    if not tables:
        ws = wb.create_sheet("数据")
        ws.cell(row=1, column=1, value="无数据")

    # 遍历每个表格，每个表格是一个工作表(Sheet)
    for i, table in enumerate(tables):
        # Excel 工作表名最长 31 字符，超长截断
        sheet_name = table.get("name", f"Sheet{i + 1}")[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = table.get("headers", [])
        rows = table.get("rows", [])

        # ============================================================
        # 样式定义
        # 为什么在循环内定义: 每个 sheet 独立，样式对象不共享
        # ============================================================
        # 表头样式: 微软雅黑(中文字体友好)、粗体、白色字、深蓝底
        h_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        # 数据样式: 微软雅黑常规
        d_font = Font(name="微软雅黑", size=10)
        # 斑马纹: 浅蓝色背景（偶数行），提升可读性
        d_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        # 细边框: 所有单元格都有
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")

        # ============================================================
        # 写入表头（第一行）
        # enumerate(headers, 1): 从 1 开始计数，因为 Excel 列从 1 开始
        # ============================================================
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = h_font        # 粗体白色字
            cell.fill = h_fill        # 深蓝背景
            cell.alignment = center   # 居中
            cell.border = thin        # 边框

        # ============================================================
        # 写入数据行（从第二行开始）
        # ============================================================
        for ri, row in enumerate(rows, 2):        # row 2 = Excel 第二行
            for ci, val in enumerate(row, 1):
                if ci <= len(headers):             # 防止数据列多于表头列
                    cell = ws.cell(row=ri, column=ci, value=_try_number(val))
                    cell.font = d_font
                    cell.border = thin
                    if ri % 2 == 0:                # 偶数行应用斑马纹
                        cell.fill = d_fill

        # ============================================================
        # 自动列宽: 根据内容最长值调整
        # 最小 8 字符，最大 60 字符（太宽影响阅读）
        # ============================================================
        for ci in range(1, len(headers) + 1):
            max_len = len(str(headers[ci - 1]))    # 表头长度
            for row in rows:
                if ci <= len(row):
                    max_len = max(max_len, len(str(row[ci - 1])))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)

    # 保存文件
    wb.save(output_path)
    return output_path


def _try_number(val):
    """
    将字符串转为数值(如果可能)

    为什么需要这个函数?
    从 CSV/文本提取的数据都是字符串，"1,234" 在 Excel 中是文本
    但用户期望它能被当作数字计算

    自动处理:
    - "1,234" → 1234  (移除千位分隔符)
    - "50.00" → 50.0
    - "ABC123" → 保持原样
    """
    if not isinstance(val, str):
        return val                      # 非字符串不处理
    cleaned = val.strip().replace(",", "")  # 清理千位分隔符
    try:
        return int(cleaned)             # 尝试整数
    except ValueError:
        try:
            return float(cleaned)       # 尝试浮点数
        except ValueError:
            return val                  # 都不是，返回原字符串


# === 注册工具 ===
registry.register(
    generate_xlsx, "generate_xlsx",
    "根据表格数据生成格式化的 XLSX 文件",
    {"tables": "表格数据列表，每项包含 name/headers/rows", "output_path": "输出路径（可选）"},
    agent="spreadsheet",
)
